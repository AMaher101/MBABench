"""Formatting tools: format_cells, freeze_panes."""
import json
from typing import Any, Dict, Optional

from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

from ..core.shared_state import mcp
from ..core.workbook_io import _get_file_path, _save_workbook_sync, _load_workbook


@mcp.tool()
async def freeze_panes(filename: str, worksheet_name: str, cell: str) -> str:
    """Freeze rows and columns at a cell position so they stay visible when scrolling.

    Freezing at "B2" keeps row 1 (headers) and column A (labels) pinned.

    Args:
        filename: Name of the Excel file
        worksheet_name: Name of the worksheet
        cell: Cell position to freeze at (e.g., "B2" freezes row 1 and column A)

    Returns:
        Success or error message
    """
    try:
        wb = _load_workbook(filename)
        ws = wb[worksheet_name]
        ws.freeze_panes = cell
        _save_workbook_sync(wb, _get_file_path(filename))
        return json.dumps({"success": True, "frozen_at": cell, "worksheet": worksheet_name})
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)})


@mcp.tool()
async def format_cells(
    filename: str,
    worksheet_name: str,
    range_address: str,
    font: Optional[Dict[str, Any]] = None,
    fill: Optional[Dict[str, Any]] = None,
    border: Optional[Dict[str, Any]] = None,
    alignment: Optional[Dict[str, Any]] = None,
    number_format: Optional[str] = None,
) -> str:
    """Apply formatting (font, fill, border, alignment, number_format) to a cell range.

    Apply ONLY after ALL calculations are verified and answer sheets linked.

    Args:
        filename: Name of the Excel file
        worksheet_name: Name of the worksheet
        range_address: Cell range (e.g., "A1:D1" for header row, "B2:B20" for data column)
        font: Font properties -- {"color": "0000FF", "bold": true, "italic": false, "size": 11, "name": "Calibri"}
        fill: Fill properties -- {"color": "002060"} (solid fill with hex color)
        border: Border properties -- {"style": "thin"} applies to all sides
        alignment: Alignment -- {"horizontal": "center", "vertical": "center", "wrap_text": true}
        number_format: Excel number format string -- "#,##0", "0.00%", "$#,##0_);($#,##0)", "0.00"

    Returns:
        Success message with count of formatted cells
    """
    try:
        wb = _load_workbook(filename)
        ws = wb[worksheet_name]

        font_obj = None
        if font:
            font_color = font.get("color")
            if isinstance(font_color, str) and font_color.startswith("#"):
                font_color = font_color[1:]
            font_obj = Font(
                color=font_color,
                bold=font.get("bold", False),
                italic=font.get("italic", False),
                size=font.get("size"),
                name=font.get("name"),
            )

        fill_obj = None
        if fill:
            fill_color = fill.get("color") or fill.get("start_color") or fill.get("fgColor") or "FFFFFF"
            if isinstance(fill_color, str) and fill_color.startswith("#"):
                fill_color = fill_color[1:]
            fill_type_val = fill.get("fill_type") or fill.get("patternType") or "solid"
            fill_obj = PatternFill(
                start_color=fill_color,
                end_color=fill_color,
                fill_type=fill_type_val,
            )

        border_obj = None
        if border:
            side = Side(style=border.get("style", "thin"))
            border_obj = Border(top=side, bottom=side, left=side, right=side)

        align_obj = None
        if alignment:
            align_obj = Alignment(
                horizontal=alignment.get("horizontal"),
                vertical=alignment.get("vertical"),
                wrap_text=alignment.get("wrap_text", False),
            )

        cell_count = 0
        from openpyxl.cell.cell import Cell as _Cell

        # Handle comma-separated non-contiguous ranges (e.g. "A1,A11,A21" or "B11:F11,B33:D33")
        sub_ranges = [r.strip() for r in range_address.split(',') if r.strip()]

        for sub_range in sub_ranges:
            target = ws[sub_range]
            if isinstance(target, _Cell):
                rows = ((target,),)
            elif isinstance(target, tuple) and target and not isinstance(target[0], tuple):
                rows = (target,)
            else:
                rows = target
            for row in rows:
                for c in row:
                    if font:
                        curr_f = c.font
                        f_color = font.get("color") or font.get("start_color")
                        if isinstance(f_color, str) and f_color.startswith("#"):
                            f_color = f_color[1:]

                        target_color = f_color if ("color" in font or "start_color" in font) else (curr_f.color if curr_f else None)
                        target_name = font.get("name", curr_f.name if curr_f else "Calibri")
                        target_size = font.get("size", curr_f.size if curr_f else 11)
                        target_bold = font.get("bold", curr_f.bold if curr_f else False)
                        target_italic = font.get("italic", curr_f.italic if curr_f else False)

                        c.font = Font(
                            name=target_name,
                            size=target_size,
                            bold=target_bold,
                            italic=target_italic,
                            color=target_color,
                        )

                    if fill:
                        curr_fill = c.fill
                        fill_color = fill.get("color") or fill.get("start_color") or fill.get("fgColor")
                        if isinstance(fill_color, str) and fill_color.startswith("#"):
                            fill_color = fill_color[1:]
                        
                        if not fill_color:
                            fill_color = (curr_fill.start_color.rgb if (curr_fill and curr_fill.start_color) else "FFFFFF")

                        fill_type_val = fill.get("fill_type") or fill.get("patternType") or (curr_fill.fill_type if curr_fill else "solid")
                        c.fill = PatternFill(
                            start_color=fill_color,
                            end_color=fill_color,
                            fill_type=fill_type_val,
                        )

                    if border:
                        curr_b = c.border
                        style = border.get("style", "thin")
                        side = Side(style=style)
                        c.border = Border(
                            top=side if ("top" in border or "style" in border) else (curr_b.top if curr_b else side),
                            bottom=side if ("bottom" in border or "style" in border) else (curr_b.bottom if curr_b else side),
                            left=side if ("left" in border or "style" in border) else (curr_b.left if curr_b else side),
                            right=side if ("right" in border or "style" in border) else (curr_b.right if curr_b else side),
                        )

                    if alignment:
                        curr_a = c.alignment
                        c.alignment = Alignment(
                            horizontal=alignment.get("horizontal", curr_a.horizontal if curr_a else None),
                            vertical=alignment.get("vertical", curr_a.vertical if curr_a else None),
                            wrap_text=alignment.get("wrap_text", curr_a.wrap_text if curr_a else False),
                        )

                    if number_format:
                        c.number_format = number_format
                    cell_count += 1

        _save_workbook_sync(wb, _get_file_path(filename))
        return json.dumps({"success": True, "cells_formatted": cell_count, "range": range_address, "worksheet": worksheet_name})
    except Exception as e:
        return json.dumps({"success": False, "error": str(e)})
