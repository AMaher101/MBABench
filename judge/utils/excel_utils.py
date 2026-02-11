import argparse
import colorsys
import csv
import io
import os
from typing import Any, Dict, List, Optional, Tuple
from zipfile import Path

import openpyxl
from logger import logger


def get_worksheet_info(workbook: openpyxl.Workbook, sheet_name: str) -> Dict[str, Any]:
    """Get high level information about a worksheet."""
    worksheet = workbook[sheet_name]
    return {
        "name": sheet_name,
        "max_row": worksheet.max_row,
        "max_column": worksheet.max_column,
        "dimensions": f"{worksheet.max_row} rows x {worksheet.max_column} columns",
    }


def load_workbooks(excel_file_path: str) -> Tuple[openpyxl.Workbook, openpyxl.Workbook]:
    """Load both formula and data-only workbooks."""
    workbook = openpyxl.load_workbook(excel_file_path, data_only=False)
    workbook_data_only = openpyxl.load_workbook(excel_file_path, data_only=True)
    return workbook, workbook_data_only


### Cell processing functions - start
def _get_formatted_value(cell, cell_data_only) -> str:
    """Get the formatted display value of a cell, preserving number formatting."""

    def _default_value(raw_value, float_rounding: int, do_rounding: bool) -> str:
        if isinstance(raw_value, float):
            if do_rounding:
                rounded_value = round(raw_value, float_rounding)
                formatted_value = f"{rounded_value}"
            else:
                formatted_value = str(raw_value)
        else:
            formatted_value = str(raw_value)
        return formatted_value

    if cell_data_only.value is None:
        return ""
    # Obtain parameters
    do_rounding = os.environ.get("DO_ROUNDING", "true").lower() == "true"
    if do_rounding:
        float_rounding = int(os.environ.get("FLOAT_ROUNDING", 6))
        percentage_rounding = int(os.environ.get("PERCENTAGE_ROUNDING", 8))

    # Get the raw value
    raw_value = cell_data_only.value

    # Check if this is a number with formatting
    if (
        isinstance(raw_value, (int, float))
        and hasattr(cell, "number_format")
        and cell.number_format
    ):
        # Get the number format from the cell
        number_format = cell.number_format

        # If it's not the default format, include format information
        if number_format != "General" and number_format != "@":

            # Get the number format from the cell
            number_format = cell.number_format
            parts = number_format.split(";")
            positive_format = parts[0] if len(parts) > 0 else None  # "$"#,##0_)
            negative_format = (
                parts[1] if len(parts) > 1 else None
            )  # [Red]\\("$"#,##0\\)
            zero_format = parts[2] if len(parts) > 2 else None  # (optional)

            if raw_value < 0 and negative_format:
                number_format = negative_format
                FORMAT_PREFIX = "NEG FORMAT"
            elif raw_value == 0 and zero_format:
                number_format = zero_format
                FORMAT_PREFIX = "ZERO FORMAT"
            else:
                number_format = positive_format
                FORMAT_PREFIX = "FORMAT"

            # For currency, percentage, and other special formats, show the format
            if "$" in number_format or "%" in number_format or "#,##0" in number_format:
                if do_rounding:
                    if "%" in number_format:
                        rounded_value = round(raw_value, percentage_rounding)
                    else:
                        rounded_value = round(
                            raw_value, float_rounding
                        )  # Round for context saving purposes
                    formatted_value = (
                        f"{rounded_value} [{FORMAT_PREFIX}:{number_format}]"
                    )
                else:
                    formatted_value = f"{raw_value} [{FORMAT_PREFIX}:{number_format}]"
        else:
            formatted_value = _default_value(raw_value, float_rounding, do_rounding)
    else:
        formatted_value = _default_value(raw_value, float_rounding, do_rounding)

    return formatted_value


def _get_excel_cell_reference(row_idx: int, col_idx: int) -> str:
    """Convert row and column indices to Excel cell reference (e.g., A1, B19)."""
    # Convert column index to Excel column letters
    col_letters = ""
    col_num = col_idx
    while col_num > 0:
        col_num -= 1
        col_letters = chr(65 + (col_num % 26)) + col_letters
        col_num //= 26

    return f"{col_letters}{row_idx}"


def _cell_has_value_or_color(cell, cell_data_only) -> bool:
    """Check if cell has value or color formatting."""
    # Check for actual value
    display_value = (
        str(cell_data_only.value) if cell_data_only.value is not None else ""
    )
    if display_value.strip():
        return True

    # Check for font color - RGB first, then indexed converted to RGB
    if cell.font and cell.font.color:
        try:
            if hasattr(cell.font.color, "rgb") and cell.font.color.rgb:
                color_value = str(cell.font.color.rgb)
                if (
                    color_value != "00000000"
                    and color_value != "Values must be of type <class 'str'>"
                ):
                    return True
            elif (
                hasattr(cell.font.color, "indexed")
                and cell.font.color.indexed is not None
            ):
                return True  # We'll convert indexed to RGB
        except (AttributeError, TypeError):
            pass

    # Check for background color - RGB first, then indexed converted to RGB
    if cell.fill:
        try:
            if hasattr(cell.fill, "start_color") and cell.fill.start_color:
                if hasattr(cell.fill.start_color, "rgb") and cell.fill.start_color.rgb:
                    bg_color = str(cell.fill.start_color.rgb)
                    if bg_color != "00000000":
                        return True
                elif (
                    hasattr(cell.fill.start_color, "indexed")
                    and cell.fill.start_color.indexed is not None
                ):
                    return True  # We'll convert indexed to RGB
        except (AttributeError, TypeError):
            pass

    return False


def _rgb_to_color_name(rgb_value: str) -> str:
    """Convert RGB value to human-readable color name using complete HSV analysis."""
    # Clean up RGB value - remove FF prefix if it's 8 characters
    if len(rgb_value) == 8 and rgb_value.startswith("FF"):
        rgb_clean = rgb_value[2:]
    elif len(rgb_value) == 8:
        # Handle ARGB format - take last 6 chars for RGB
        rgb_clean = rgb_value[2:]
    else:
        rgb_clean = rgb_value

    # Ensure we have exactly 6 characters
    if len(rgb_clean) != 6:
        return rgb_clean

    try:
        # Parse RGB values
        r = int(rgb_clean[0:2], 16)
        g = int(rgb_clean[2:4], 16)
        b = int(rgb_clean[4:6], 16)

        # Convert to HSV using colorsys
        h, s, v = colorsys.rgb_to_hsv(r / 255, g / 255, b / 255)
        h = h * 360  # Convert to degrees

        # First check saturation and value for achromatic colors
        if s < 0.1:
            if v < 0.2:
                color_name = "black"
            elif v > 0.9:
                color_name = "white"
            else:
                color_name = "gray"

        # Low value = dark colors
        elif v < 0.4:  # Increased threshold for better dark color detection
            if h < 30 or h >= 330:
                color_name = "dark_red"
            elif h < 75:  # Extended range for brown detection
                color_name = "brown"
            elif h < 150:
                color_name = "dark_green"
            elif h < 270:
                color_name = "dark_blue"
            else:
                color_name = "dark_purple"

        # Low saturation but not achromatic = muted colors
        elif s < 0.4:  # Slightly increased threshold
            if h < 60:
                color_name = "beige"
            elif h < 150:
                color_name = "olive"
            elif h < 210:
                color_name = "slate_blue"
            else:
                color_name = "muted_purple"

        # High saturation and value = bright colors
        elif s > 0.7 and v > 0.7:
            if h < 15 or h >= 345:
                color_name = "bright_red"
            elif h < 45:
                color_name = "bright_orange"
            elif h < 65:
                color_name = "bright_yellow"
            elif h < 170:
                color_name = "bright_green"
            elif h < 200:
                color_name = "bright_cyan"
            elif h < 260:
                color_name = "bright_blue"
            elif h < 310:
                color_name = "bright_purple"
            else:
                color_name = "bright_magenta"

        # Medium to high saturation with high value = pastel colors
        elif s < 0.7 and v > 0.7:  # Better pastel detection
            if h < 15 or h >= 330:
                color_name = "pink"
            elif h < 45:
                color_name = "peach"
            elif h < 65:
                color_name = "pale_yellow"
            elif h < 170:
                color_name = "pale_green"
            elif h < 200:
                color_name = "pale_cyan"
            elif h < 260:
                color_name = "pale_blue"
            elif h < 310:
                color_name = "lavender"
            else:
                color_name = "pale_magenta"

        # Finally use hue for normal saturated colors
        else:
            if h < 15 or h >= 345:
                color_name = "red"
            elif h < 45:
                color_name = "orange"
            elif h < 65:
                color_name = "yellow"
            elif h < 170:
                color_name = "green"
            elif h < 200:
                color_name = "cyan"
            elif h < 260:
                color_name = "blue"
            elif h < 310:
                color_name = "purple"
            else:
                color_name = "magenta"

        return f"{color_name} ({rgb_clean})"

    except ValueError:
        # Fallback to just returning the hex code if parsing fails
        return rgb_clean


def _convert_indexed_to_rgb(indexed_val: int) -> str:
    """Convert indexed color to RGB with custom overrides for specific indices."""
    # Custom indexed color overrides
    CUSTOM_INDEXED_COLORS = {
        13: "00FF00",  # Override index 13 (was FFFF00/yellow) -> green for A26
        14: "FF6600",  # Override index 14 (was FF00FF/magenta) -> orange for A27
    }

    # Check for custom override first
    if indexed_val in CUSTOM_INDEXED_COLORS:
        return CUSTOM_INDEXED_COLORS[indexed_val]

    # Otherwise use standard COLOR_INDEX
    from openpyxl.styles.colors import COLOR_INDEX

    if indexed_val < len(COLOR_INDEX):
        argb_hex = COLOR_INDEX[indexed_val]
        if len(argb_hex) == 8:
            return argb_hex[2:]  # Skip alpha (AA) part, take RRGGBB

    return "000000"  # Default to black if conversion fails


def extract_cell_formatting(cell) -> str:
    """Extract formatting information from a cell."""
    formatting_parts = []

    try:
        # Font formatting
        if cell.font:
            font = cell.font
            if font.name and font.name != "Calibri":  # Skip default font
                formatting_parts.append(f"font:{font.name}")
            if font.size and font.size != 11:  # Skip default size
                formatting_parts.append(f"size:{font.size}")
            if font.bold:
                formatting_parts.append("bold")
            if font.italic:
                formatting_parts.append("italic")
            if font.underline:
                formatting_parts.append("underline")

            # Handle font color safely with explicit textcolor label and color names - RGB FIRST
            if font.color:
                try:
                    color_added = False

                    # RGB first priority - direct RGB colors
                    if hasattr(font.color, "rgb") and font.color.rgb:
                        color_value = str(font.color.rgb)
                        # Only use RGB if it's actually valid (not error or default)
                        if (
                            color_value != "00000000"
                            and color_value != "None"
                            and "Values must be of type" not in color_value
                        ):
                            color_name = _rgb_to_color_name(color_value)
                            formatting_parts.append(f"textcolor:{color_name}")
                            color_added = True

                    # If no RGB available or RGB invalid, convert indexed to RGB first, then apply HSV
                    if (
                        not color_added
                        and hasattr(font.color, "indexed")
                        and font.color.indexed is not None
                    ):
                        try:
                            indexed_val = int(
                                font.color.indexed
                            )  # Convert to int to handle openpyxl Integer objects
                            # Use custom conversion function with overrides
                            rgb_hex = _convert_indexed_to_rgb(indexed_val)
                            # Now apply HSV analysis to the converted RGB
                            color_name = _rgb_to_color_name(rgb_hex)
                            formatting_parts.append(f"textcolor:{color_name}")
                        except (ValueError, TypeError):
                            pass  # Skip if conversion fails

                except (AttributeError, TypeError):
                    pass  # Skip problematic color values

        # Fill/background color with explicit bgcolor label and color names - RGB FIRST
        if cell.fill:
            try:
                if hasattr(cell.fill, "start_color") and cell.fill.start_color:
                    bg_color_added = False

                    # RGB first priority - direct RGB colors
                    if (
                        hasattr(cell.fill.start_color, "rgb")
                        and cell.fill.start_color.rgb
                    ):
                        bg_color = str(cell.fill.start_color.rgb)
                        # Only use RGB if it's actually valid (not error or default)
                        if (
                            bg_color != "00000000"
                            and "Values must be of type" not in bg_color
                        ):
                            color_name = _rgb_to_color_name(bg_color)
                            formatting_parts.append(f"bgcolor:{color_name}")
                            bg_color_added = True

                    # If no RGB available or RGB invalid, convert indexed to RGB first, then apply HSV
                    if not bg_color_added and (
                        hasattr(cell.fill.start_color, "indexed")
                        and cell.fill.start_color.indexed is not None
                    ):
                        try:
                            indexed_val = int(
                                cell.fill.start_color.indexed
                            )  # Convert to int to handle openpyxl Integer objects
                            # Use custom conversion function with overrides
                            rgb_hex = _convert_indexed_to_rgb(indexed_val)
                            # Now apply HSV analysis to the converted RGB
                            color_name = _rgb_to_color_name(rgb_hex)
                            formatting_parts.append(f"bgcolor:{color_name}")
                        except (ValueError, TypeError):
                            pass  # Skip if conversion fails
            except (AttributeError, TypeError):
                pass  # Skip problematic fill values

        # Alignment (removed wrap_text to exclude 'wrap' from formatting)
        if cell.alignment:
            align = cell.alignment
            if align.horizontal and align.horizontal != "general":
                formatting_parts.append(f"halign:{align.horizontal}")
            if align.vertical and align.vertical != "bottom":
                formatting_parts.append(f"valign:{align.vertical}")
            # Removed: if align.wrap_text: formatting_parts.append("wrap")

        # Borders
        if cell.border:
            border_parts = []
            for side in ["top", "bottom", "left", "right"]:
                try:
                    border_side = getattr(cell.border, side)
                    if (
                        border_side
                        and hasattr(border_side, "style")
                        and border_side.style
                    ):
                        border_parts.append(f"{side}:{border_side.style}")
                except (AttributeError, TypeError):
                    continue
            if border_parts:
                formatting_parts.append(f"border:{','.join(border_parts)}")

    except Exception as e:
        # If any error occurs, return empty formatting rather than crash
        logger.info(f"Warning: Error extracting formatting for cell: {e}")
        return ""

    return ";".join(formatting_parts)


def create_enhanced_cell(
    cell,
    cell_data_only,
    row_idx: int,
    col_idx: int,
) -> str:
    """Create enhanced cell data with embedded formula and formatting."""
    # Start with formatted display value (includes number formatting)
    display_value = _get_formatted_value(cell, cell_data_only)

    # For non-empty cells, start with cell reference
    if display_value.strip():
        cell_ref = _get_excel_cell_reference(row_idx, col_idx)
        cell_parts = [f"[{cell_ref}]{display_value}"]
    else:
        cell_parts = [display_value]

    # Add formula if present
    if isinstance(cell.value, str) and cell.value.startswith("="):
        cell_parts.append(f"FORMULA:{cell.value}")

    # Add formatting if present and cell has value or color formatting
    if _cell_has_value_or_color(cell, cell_data_only):
        formatting = extract_cell_formatting(cell)
        if formatting:
            cell_parts.append(f"FORMAT:{formatting}")

    return "|".join(cell_parts)


def list_to_csv(data: List[List[str]]) -> str:
    """Convert list of lists to CSV string."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerows(data)
    return output.getvalue()


def extract_all_cell_data(worksheet, worksheet_data) -> Dict[str, Any]:
    """Extract everything into one CSV."""
    enhanced_data = []

    for row_idx, row in enumerate(worksheet.iter_rows(), 1):
        enhanced_row = []
        for col_idx, cell in enumerate(row, 1):
            cell_data_only = worksheet_data.cell(row_idx, col_idx)
            enhanced_cell = create_enhanced_cell(
                cell,
                cell_data_only,
                row_idx,
                col_idx,
            )
            enhanced_row.append(enhanced_cell)
        enhanced_data.append(enhanced_row)

    return {
        "full": list_to_csv(enhanced_data),
        "metadata": {
            "format": "Enhanced with cell references, formulas and formatting",
            "separator": "|",
            "rows": len(enhanced_data),
            "columns": len(enhanced_data[0]) if enhanced_data else 0,
        },
    }


### Cell processing functions - end
### Additional Info Extraction functions - start
def create_safe_filename(sheet_name: str) -> str:
    """Create a safe filename from sheet name."""
    return "".join(c for c in sheet_name if c.isalnum() or c in (" ", "-", "_")).strip()


def extract_merged_cells_info(worksheet) -> List[str]:
    """Extract merged cell ranges from a worksheet."""
    merged_cells = []
    for merged_range in worksheet.merged_cells.ranges:
        # Convert merged range to string format like "A1:C3"
        merged_cells.append(str(merged_range))
    return merged_cells


def extract_frozen_panes_info(worksheet) -> Dict[str, Any]:
    """Extract frozen panes information from a worksheet."""
    frozen_info = {"has_frozen_panes": False, "freeze_panes": None, "split_panes": None}

    try:
        # Check for frozen panes
        if hasattr(worksheet, "freeze_panes") and worksheet.freeze_panes:
            frozen_info["has_frozen_panes"] = True
            freeze_cell = worksheet.freeze_panes
            if freeze_cell:
                frozen_info["freeze_panes"] = str(freeze_cell)

        # Check for split panes (less common but possible)
        if hasattr(worksheet, "sheet_view") and worksheet.sheet_view:
            sheet_view = worksheet.sheet_view
            if hasattr(sheet_view, "pane") and sheet_view.pane:
                pane = sheet_view.pane
                if pane.xSplit or pane.ySplit:
                    frozen_info["split_panes"] = {
                        "xSplit": pane.xSplit,
                        "ySplit": pane.ySplit,
                        "topLeftCell": pane.topLeftCell,
                        "activePane": pane.activePane,
                        "state": pane.state,
                    }
    except (AttributeError, TypeError):
        # Handle cases where properties don't exist or have unexpected types
        pass

    return frozen_info


### Additional Info Extraction functions - end


def save_additional_format_info(
    output_dir: Path, sheet_name: str, worksheet, quiet: bool = False
) -> str:
    """Save additional formatting information (merged cells, frozen panes) to a text file."""
    output_dir.mkdir(exist_ok=True)
    safe_sheet_name = create_safe_filename(sheet_name)

    # Create the additional format file
    format_file_path = output_dir / f"{safe_sheet_name}_additional_format.txt"

    # Extract information
    merged_cells = extract_merged_cells_info(worksheet)
    frozen_panes = extract_frozen_panes_info(worksheet)

    # Create content
    content_lines = []
    content_lines.append(f"Additional Format Information for Sheet: {sheet_name}")
    content_lines.append("=" * 60)
    content_lines.append("")

    # Merged Cells Section
    content_lines.append("MERGED CELLS:")
    content_lines.append("-" * 20)
    if merged_cells:
        for i, merged_range in enumerate(merged_cells, 1):
            content_lines.append(f"{i}. {merged_range}")
    else:
        content_lines.append("No merged cells found.")
    content_lines.append("")

    # Frozen Panes Section
    content_lines.append("FROZEN PANES:")
    content_lines.append("-" * 20)
    if frozen_panes["has_frozen_panes"]:
        if frozen_panes["freeze_panes"]:
            content_lines.append(f"Freeze Panes: {frozen_panes['freeze_panes']}")
        if frozen_panes["split_panes"]:
            content_lines.append("Split Panes Information:")
            split_info = frozen_panes["split_panes"]
            for key, value in split_info.items():
                if value is not None:
                    content_lines.append(f"  {key}: {value}")
    else:
        content_lines.append("No frozen panes found.")
    content_lines.append("")

    # Metadata
    content_lines.append("METADATA:")
    content_lines.append("-" * 20)
    content_lines.append(
        f"Sheet Dimensions: {worksheet.max_row} rows x {worksheet.max_column} columns"
    )
    content_lines.append(f"Total Merged Cell Ranges: {len(merged_cells)}")
    content_lines.append(f"Has Frozen Panes: {frozen_panes['has_frozen_panes']}")

    # Write to file
    content = "\n".join(content_lines)
    format_file_path.write_text(content, encoding="utf-8")

    if not quiet:
        logger.info(f"Saved additional format info: {format_file_path}")

    return str(format_file_path)


def save_sheet_csv_files(
    output_dir: Path,
    sheet_name: str,
    extraction_result: Dict[str, Any],
    worksheet=None,
    quiet: bool = False,
) -> List[str]:
    """Save CSV files and additional format info for a single sheet to directory."""
    output_dir.mkdir(exist_ok=True)
    safe_sheet_name = create_safe_filename(sheet_name)
    saved_files = []

    # Save based on extraction result content

    if "full" in extraction_result:
        enhanced_path = output_dir / f"{safe_sheet_name}_full.csv"
        enhanced_path.write_text(extraction_result["full"], encoding="utf-8")
        saved_files.append(str(enhanced_path))
        if not quiet:
            logger.info(f"Saved: {enhanced_path}")

    # Save additional format information (merged cells, frozen panes)
    if worksheet is not None:
        additional_format_path = save_additional_format_info(
            output_dir, sheet_name, worksheet, quiet=quiet
        )
        saved_files.append(additional_format_path)

    return saved_files


def process_all_worksheets(
    excel_file_path: str,
    output_dir: Path,
    quiet: bool = False,
    run_calculation: bool = False,
) -> Dict[str, Any]:
    """
    Process all worksheets in an Excel file and extract to CSV.

    Args:
        excel_file_path: Path to Excel file
        output_dir: Directory to save CSV files

    Returns:
        Dictionary with results for each sheet
    """

    def recalculate_xlsx(filepath: str, outdir: str = "."):
        """Re-save xlsx through LibreOffice to trigger formula calculation."""
        import subprocess

        subprocess.run(
            [
                os.getenv("LIBREOFFICE_PATH", "libreoffice"),
                "--headless",
                "--calc",
                "--convert-to",
                "xlsx",
                "--outdir",
                outdir,
                filepath,
            ],
            check=True,
            cwd=os.getcwd(),
        )

    if run_calculation:
        # Sometimes. Excel files created by openpyxl won't calculate the cached values from the formulas. This process triggers a recalculation through Libreoffice.
        # Make recalulated file under same directory
        temp_output_dir = Path(excel_file_path).parent / "temp_recalculated"
        temp_output_dir.mkdir(exist_ok=True)
        logger.info(
            f"INFO: Recalculating Excel file using LibreOffice: {excel_file_path}. Putting recalculated file in {temp_output_dir}"
        )
        recalculate_xlsx(excel_file_path, str(temp_output_dir))
        recalculated_file = temp_output_dir / Path(excel_file_path).name
        if recalculated_file.exists():
            excel_file_path = str(recalculated_file)
        else:
            raise FileNotFoundError(
                f"Recalculated file not found: {recalculated_file}. Something may have gone wrong with LibreOffice conversion."
            )
    if not output_dir.exists():
        output_dir.mkdir(parents=True, exist_ok=True)

    # Load workbooks
    workbook, workbook_data_only = load_workbooks(excel_file_path)

    # Get all sheet names
    sheet_names = workbook.sheetnames
    if not quiet:
        logger.info(f"Found {len(sheet_names)} worksheets: {sheet_names}")

    results = {}
    all_saved_files = []

    # Process each worksheet
    for sheet_name in sheet_names:
        if not quiet:
            logger.info(f"\n=== Processing sheet: {sheet_name} ===")

        # Get worksheets
        worksheet = workbook[sheet_name]
        worksheet_data = workbook_data_only[sheet_name]

        # Get sheet info
        if type(worksheet) is openpyxl.chartsheet.chartsheet.Chartsheet:
            logger.info(f"Skipping ChartSheet: {sheet_name}")
            continue

        sheet_info = get_worksheet_info(workbook, sheet_name)
        if not quiet:
            logger.info(f"Dimensions: {sheet_info['dimensions']}")

        # Extract data
        extraction_result = extract_all_cell_data(worksheet, worksheet_data)

        # Save files
        saved_files = save_sheet_csv_files(
            output_dir, sheet_name, extraction_result, worksheet, quiet=quiet
        )

        # Store results
        results[sheet_name] = {
            "info": sheet_info,
            "extraction": extraction_result,
            "saved_files": saved_files,
        }

        all_saved_files.extend(saved_files)

    # Summary
    if not quiet:
        logger.info(f"\n=== Summary ===")
        logger.info(f"Processed {len(sheet_names)} worksheets")
        logger.info(f"Generated {len(all_saved_files)} CSV files")
        logger.info(f"Output directory: {output_dir.absolute()}")

    return {
        "sheets": results,
        "summary": {
            "total_sheets": len(sheet_names),
            "total_files": len(all_saved_files),
            "output_directory": str(output_dir.absolute()),
            "all_files": all_saved_files,
        },
    }


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Test csv extraction.")
    parser.add_argument(
        "--filename",
        "-f",
        nargs="?",
        default="test.xlsx",
        help="Excel file to process (default: test.xlsx)",
    )

    parser.add_argument(
        "--output-dir",
        "-o",
        default=None,
        help="Output directory for CSV files (default: ./extracted_csvs)",
    )
    parser.add_argument(
        "--run-recalc",
        action="store_true",
        help="Recalculate formulas using LibreOffice before extraction",
    )

    args = parser.parse_args()
    if args.output_dir is None:
        from misc_utils import load_project_configs, relative_path_from_project_root

        configs = load_project_configs()
        args.output_dir = (
            configs.get("SCRATCH_PATH", "./scratch")
            + f"/extracted_csvs_output/{Path(args.filename).stem}"
        )
        args.output_dir = relative_path_from_project_root(args.output_dir)
        print(
            f"Output directory not provided. Using {args.output_dir} based on SCRATCH_PATH from project configs."
        )

    process_all_worksheets(
        excel_file_path=args.filename,
        output_dir=args.output_dir,
        quiet=False,
        run_calculation=args.run_recalc,
    )
